import io
import re
import json
import hashlib
import threading
from pathlib import Path
from datetime import datetime

import fitz
import pandas as pd
import ollama
from PIL import Image as PILImage
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── CONFIG ───────────────────────────────────────────────────────────────
OCR_MODEL    = "glm-ocr-hires:latest"      # purpose-built OCR — reads documents perfectly
TEXT_MODEL   = "deepseek-r1:8b"     # reasoning — reconstructs table structure
OUTPUT_DIR   = Path("./excel_exports")

MIN_ROWS         = 1
MIN_IMAGE_WIDTH  = 150
MIN_IMAGE_HEIGHT = 150
VISION_TIMEOUT   = 200              # seconds before giving up on a single image
OCR_MAX_WIDTH    = 800              # resize images before OCR to save VRAM

JUNK_PHRASES = [
    "here is the requested image",
    "this is the table",
    "<!-- image -->",
]

HEADER_BG_COLOR   = "1F4E79"
HEADER_FONT_COLOR = "FFFFFF"
ALT_ROW_COLOR     = "D6E4F0"
BORDER_COLOR      = "BDD7EE"
SUMMARY_BG_COLOR  = "E2EFDA"


# ── HELPERS ──────────────────────────────────────────────────────────────

def strip_thinking(text: str) -> str:
    return re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL).strip()


def make_border():
    side = Side(border_style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def resize_image_for_ocr(image_bytes: bytes, max_width: int = OCR_MAX_WIDTH) -> bytes:
    """
    Resize image before sending to OCR model.
    800px wide is sufficient for GLM-OCR to read text accurately.
    Reduces VRAM usage and inference time significantly.
    """
    img  = PILImage.open(io.BytesIO(image_bytes))
    w, h = img.size

    if w <= max_width:
        return image_bytes

    scale = max_width / w
    img   = img.resize((max_width, int(h * scale)), PILImage.LANCZOS)
    buf   = io.BytesIO()
    img.save(buf, format="PNG")
    print(f"       Resized: {w}x{h} → {max_width}x{int(h * scale)}px")
    return buf.getvalue()


def call_with_timeout(fn, timeout: int):
    """
    Run fn() in a thread with a timeout.
    Returns (result, error) — one will always be None.
    """
    holder = {"result": None, "error": None}

    def _run():
        try:
            holder["result"] = fn()
        except Exception as e:
            holder["error"] = str(e)

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    t.join(timeout=timeout)

    if t.is_alive():
        return None, f"Timed out after {timeout}s"

    return holder["result"], holder["error"]


# ── MARKDOWN → DATAFRAME ─────────────────────────────────────────────────

def markdown_to_dataframe(markdown: str) -> pd.DataFrame | None:
    """
    Parse a markdown table string into a pandas DataFrame.
    Handles malformed markdown gracefully.
    Falls back to JSON if markdown parsing fails.
    """
    # ── Attempt 1: manual pipe parsing (most reliable) ────────────────
    try:
        lines = [
            l.strip() for l in markdown.splitlines()
            if l.strip().startswith("|") and l.strip().endswith("|")
            and not re.match(r"^\|[-:| ]+\|$", l.strip())
        ]
        if len(lines) >= 2:
            headers = [h.strip() for h in lines[0].split("|") if h.strip()]
            rows    = []
            for line in lines[1:]:
                cells = [c.strip() for c in line.split("|") if c.strip() != ""]
                while len(cells) < len(headers):
                    cells.append("")
                rows.append(cells[:len(headers)])

            df = pd.DataFrame(rows, columns=headers)
            for col in df.columns:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace({"nan": "", "None": ""})

            df = df.dropna(how="all").reset_index(drop=True)
            df = df[df.apply(lambda r: any(str(v).strip() for v in r), axis=1)]

            if not df.empty:
                return df
    except Exception:
        pass

    # ── Attempt 2: JSON fallback ──────────────────────────────────────
    try:
        json_match = re.search(r"\[.*\]", markdown, re.DOTALL)
        if json_match:
            data = json.loads(json_match.group())
            if isinstance(data, list) and data:
                df = pd.DataFrame(data).dropna(how="all").reset_index(drop=True)
                if not df.empty:
                    return df
    except Exception:
        pass

    return None


# ── TWO-PASS IMAGE TABLE EXTRACTION ──────────────────────────────────────

def extract_complex_image_table(
    image_bytes: bytes,
    page_num: int,
    img_idx: int
) -> list[dict]:
    """
    Two-pass extraction:
    Pass A — GLM-OCR reads raw text from image
    Pass B — deepseek-r1: outputs JSON with strict schema validation
    """
    results = []

    # ── PASS A: GLM-OCR ───────────────────────────────────────────────
    print(f"       Pass A: OCR with {OCR_MODEL}...")

    def run_ocr():
        return ollama.chat(
            model=OCR_MODEL,
            messages=[{
                "role":    "user",
                "content": "Text Recognition:",
                "images":  [image_bytes]
            }]
        ).message.content.strip()

    raw_text, err = call_with_timeout(run_ocr, timeout=VISION_TIMEOUT)

    if err or not raw_text or len(raw_text) < 20:
        print(f"       ⚠️  Pass A failed: {err or 'empty response'}")
        return []

    print(f"       Pass A complete: {len(raw_text)} chars")
    print(f"       Preview: {raw_text}...")

    # ── PASS B: deepseek-r1 ───────────────────────────────────────────
    print(f"       Pass B: Structuring with {TEXT_MODEL}...")

    # Define the exact schema we want
    schema = {
        "type": "object",
        "properties": {
            "tables": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "name": {
                            "type": "string",
                            "description": "Short table name (max 4 words)"
                        },
                        "headers": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "Column names"
                        },
                        "rows": {
                            "type": "array",
                            "items": {
                                "type": "array",
                                "items": {
                                    "oneOf": [
                                        {"type": "string"},
                                        {"type": "number"},
                                        {"type": "null"}
                                    ]
                                }
                            },
                            "description": "Each row is an array matching headers length"
                        }
                    },
                    "required": ["name", "headers", "rows"]
                }
            }
        },
        "required": ["tables"]
    }

    def run_structure():
        return ollama.chat(
            model=TEXT_MODEL,
            messages=[{
                "role": "user",
                "content": f"""The following text was OCR'd from an image containing tables.

Convert ALL tables into the JSON format specified below.

Rules:
- Each row array must have EXACTLY the same length as headers
- For merged/spanning cells, repeat the value on every row it spans
- Preserve numbers as numbers, not strings (e.g. 100 not "100")
- Use null for empty cells
- Strip logos, headers, footers, page numbers, run dates
- Table names: max 4 words, descriptive

JSON Schema (you MUST output valid JSON matching this schema):
{json.dumps(schema, indent=2)}

OCR TEXT:
{raw_text}

Output ONLY valid JSON. No markdown code blocks, no explanation."""
            }],
            format="json"  # ✅ Ollama's constrained decoding — guarantees valid JSON
        ).message.content

    raw_response, err = call_with_timeout(run_structure, timeout=120)

    if err:
        print(f"       ⚠️  Pass B failed: {err}")
        return []

    # ── PARSE: JSON is guaranteed valid by format="json" ──────────────
    try:
        data = json.loads(strip_thinking(raw_response))
        tables_json = data.get("tables", [])

        if not tables_json:
            print(f"       ⚠️  No tables in JSON response")
            return []

        print(f"       {len(tables_json)} table(s) in JSON")

        for tbl_data in tables_json:
            name    = tbl_data.get("name", "Untitled")
            headers = tbl_data.get("headers", [])
            rows    = tbl_data.get("rows", [])

            if not headers or not rows:
                continue

            # ✅ Build DataFrame directly from structured data
            # No markdown parsing, no regex, no ambiguity
            df = pd.DataFrame(rows, columns=headers)

            # Clean
            for col in df.columns:
                if df[col].dtype == object:
                    df[col] = df[col].astype(str).str.strip()
                    df[col] = df[col].replace({"nan": "", "None": "", "null": ""})

            df = df.dropna(how="all").reset_index(drop=True)
            df = df[df.apply(lambda r: any(str(v).strip() for v in r), axis=1)]

            if df.empty or len(df) < MIN_ROWS:
                continue

            safe_name = re.sub(r'[\\/*?:\[\]]', '', name)[:25]
            label     = f"P{page_num+1}_{safe_name}"[:31]

            results.append({
                "df":        df,
                "page":      page_num + 1,
                "table_num": img_idx + 1,
                "label":     label,
                "rows":      len(df),
                "cols":      len(df.columns),
                "source":    "image"
            })
            print(f"       ✅ '{name}': {len(df)} rows × {len(df.columns)} cols")

    except json.JSONDecodeError as e:
        print(f"       ⚠️  JSON parse failed: {e}")
        print(f"       Response: {raw_response[:300]}")
    except Exception as e:
        print(f"       ⚠️  DataFrame build failed: {e}")

    return results

# ── PAGE EXTRACTION ───────────────────────────────────────────────────────

def extract_image_tables_from_page(
    page: fitz.Page,
    page_num: int,
    doc: fitz.Document,
    native_hashes: set
) -> list[dict]:
    """
    Extract tables from images embedded on a page.
    Uses two-pass GLM-OCR + deepseek-r1 approach.
    Skips images already processed as native tables.
    """
    image_tables = []
    image_list   = page.get_images(full=True)
    seen_hashes  = set(native_hashes)

    for img_idx, img_info in enumerate(image_list):
        try:
            xref        = img_info[0]
            base_image  = doc.extract_image(xref)
            image_bytes = base_image["image"]
            width       = base_image.get("width",  0)
            height      = base_image.get("height", 0)

            if width < MIN_IMAGE_WIDTH or height < MIN_IMAGE_HEIGHT:
                continue

            img_hash = hashlib.md5(image_bytes).hexdigest()
            if img_hash in seen_hashes:
                print(f"     → Image {img_idx + 1}: already processed, skipping")
                continue
            seen_hashes.add(img_hash)

            print(f"     → Image {img_idx + 1} ({width}x{height}px)")

            # Resize before OCR
            image_bytes = resize_image_for_ocr(image_bytes)

            # Two-pass extraction
            extracted = extract_complex_image_table(image_bytes, page_num, img_idx)
            image_tables.extend(extracted)

        except Exception as e:
            print(f"     ⚠️  Image {img_idx + 1} on page {page_num + 1}: {e}")

    return image_tables


def extract_tables_from_pdf(pdf_path: str) -> list[dict]:
    """
    Full two-pass extraction across all pages.
    Pass 1: PyMuPDF native table detection (instant)
    Pass 2: GLM-OCR + deepseek-r1 on embedded images (GPU)
    """
    fitz_doc   = fitz.open(pdf_path)
    all_tables = []

    print(f"\n📄 Scanning {len(fitz_doc)} page(s)...")
    print(f"   Pass 1: Native PDF tables  (PyMuPDF — instant)")
    print(f"   Pass 2: Image-based tables (GLM-OCR → deepseek-r1)\n")

    for page_num in range(len(fitz_doc)):
        page = fitz_doc[page_num]
        print(f"  ── Page {page_num + 1} {'─'*40}")

        native_hashes = set()

        # ── Pass 1: Native ────────────────────────────────────────────
        try:
            detected = page.find_tables()
            if detected and detected.tables:
                print(f"  Pass 1: {len(detected.tables)} native table(s)")
                for tbl_idx, table in enumerate(detected.tables):
                    try:
                        df = table.to_pandas()
                        if df.empty:
                            continue

                        df.columns = [str(c) for c in df.columns]
                        junk_cols  = [
                            c for c in df.columns
                            if any(j in str(c).lower() for j in JUNK_PHRASES)
                            and len(str(c)) > 10
                        ]
                        df = df.drop(columns=junk_cols, errors="ignore")
                        df = df.dropna(how="all")
                        df = df.loc[:, df.notna().any()]

                        for col in df.columns:
                            if df[col].dtype == object:
                                df[col] = df[col].astype(str).str.strip()
                                df[col] = df[col].replace("nan", "")

                        headers_generated = all(
                            str(c).startswith("Col") or str(c).isdigit()
                            for c in df.columns
                        )
                        if headers_generated and len(df) > 0:
                            df.columns = [str(v) for v in df.iloc[0].tolist()]
                            df = df.iloc[1:].reset_index(drop=True)

                        if len(df) < MIN_ROWS:
                            continue

                        all_tables.append({
                            "df":        df,
                            "page":      page_num + 1,
                            "table_num": tbl_idx + 1,
                            "label":     f"Page{page_num+1}_Table{tbl_idx+1}",
                            "rows":      len(df),
                            "cols":      len(df.columns),
                            "source":    "native"
                        })
                        print(f"     → Native table {tbl_idx+1}: "
                              f"{len(df)} rows × {len(df.columns)} cols")

                    except Exception as e:
                        print(f"     ⚠️  Native table {tbl_idx+1}: {e}")
            else:
                print(f"  Pass 1: No native tables")
        except Exception as e:
            print(f"  Pass 1 failed: {e}")

        # ── Pass 2: Images ────────────────────────────────────────────
        image_list = page.get_images(full=True)
        if image_list:
            print(f"  Pass 2: {len(image_list)} image(s) on page")
            image_tables = extract_image_tables_from_page(
                page, page_num, fitz_doc, native_hashes
            )
            all_tables.extend(image_tables)
        else:
            print(f"  Pass 2: No images on page")

    fitz_doc.close()

    native_n = sum(1 for t in all_tables if t.get("source") == "native")
    image_n  = sum(1 for t in all_tables if t.get("source") == "image")

    print(f"\n{'='*52}")
    print(f"  Total tables: {len(all_tables)}")
    print(f"    Native:      {native_n}")
    print(f"    Image-based: {image_n}")
    print(f"{'='*52}")

    return all_tables


# ── EXCEL EXPORT ──────────────────────────────────────────────────────────

def write_table_to_sheet(ws, df: pd.DataFrame, table_label: str):
    header_font  = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    header_fill  = PatternFill("solid", fgColor=HEADER_BG_COLOR)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alt_fill     = PatternFill("solid", fgColor=ALT_ROW_COLOR)
    center_align = Alignment(horizontal="center", vertical="center")
    border       = make_border()

    for col_idx, col_name in enumerate(df.columns, 1):
        cell           = ws.cell(row=1, column=col_idx, value=str(col_name))
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, value in enumerate(row, 1):
            try:
                value = int(value) if str(value).isdigit() \
                    else float(value) \
                    if str(value).replace(".", "").replace(
                        "-", "").replace("$", "").isdigit() \
                    else value
            except (ValueError, TypeError):
                pass
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center_align
            cell.border    = border
            if is_alt:
                cell.fill  = alt_fill

    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        col_values = [str(v) for v in df.iloc[:, col_idx - 1]]
        max_length = max(len(str(df.columns[col_idx - 1])), *[len(v) for v in col_values])
        ws.column_dimensions[col_letter].width = max(10, min(50, max_length + 4))

    ws.freeze_panes             = "A2"
    ws.row_dimensions[1].height = 25

    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        col_data   = df.iloc[:, col_idx - 1]
        numeric    = pd.to_numeric(
            col_data.astype(str).str.replace("$", "").str.replace(",", ""),
            errors="coerce"
        )
        if numeric.notna().any():
            last_row = len(df) + 1
            ws.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{last_row}",
                ColorScaleRule(
                    start_type="min",      start_color="F8696B",
                    mid_type="percentile", mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",        end_color="63BE7B"
                )
            )


def export_to_excel(tables: list[dict], pdf_path: str) -> Path:
    OUTPUT_DIR.mkdir(exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = OUTPUT_DIR / f"{Path(pdf_path).stem}_tables_{timestamp}.xlsx"

    print(f"\n📊 Exporting {len(tables)} table(s)...")

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        contents_df = pd.DataFrame({
            "Sheet":   [t["label"]                for t in tables],
            "Page":    [t["page"]                 for t in tables],
            "Rows":    [t["rows"]                 for t in tables],
            "Columns": [t["cols"]                 for t in tables],
            "Source":  [t.get("source", "native") for t in tables],
        })
        contents_df.to_excel(writer, sheet_name="Contents", index=False)
        for tbl in tables:
            tbl["df"].to_excel(writer, sheet_name=tbl["label"], index=False)

    wb = load_workbook(out_path)
    write_table_to_sheet(wb["Contents"], contents_df, "Contents")
    for tbl in tables:
        write_table_to_sheet(wb[tbl["label"]], tbl["df"], tbl["label"])
    wb.save(out_path)

    print(f"  ✅ Saved: {out_path}")
    return out_path


# ── LLM ANALYSIS ──────────────────────────────────────────────────────────

def analyse_tables(tables: list[dict], out_path: Path) -> None:
    print(f"\n🔍 Analysing {len(tables)} table(s) with {TEXT_MODEL}...")

    wb = load_workbook(out_path)
    if "Analysis" in wb.sheetnames:
        del wb["Analysis"]

    ws = wb.create_sheet("Analysis", 0)
    ws["A1"]           = "AI Analysis Report"
    ws["A1"].font      = Font(bold=True, size=16, color=HEADER_FONT_COLOR)
    ws["A1"].fill      = PatternFill("solid", fgColor=HEADER_BG_COLOR)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height     = 35
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 90
    ws["A2"] = "Generated"
    ws["B2"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A3"] = "Source"
    ws["B3"] = str(out_path)

    current_row = 5

    for tbl in tables:
        label  = tbl["label"]
        source = tbl.get("source", "native")
        print(f"  → '{label}' ({source})...")

        response = ollama.chat(
            model=TEXT_MODEL,
            messages=[{
                "role": "user",
                "content": f"""Analyse this data table:

Table: {label} | Source: {source}
{tbl['df'].to_markdown(index=False)}

Provide:
1. 2-3 sentence summary of what this table represents
2. Key observations (patterns, outliers, notable values)
3. Data quality issues (missing values, inconsistencies)
4. Two specific business insights

Be concise and specific."""
            }]
        )
        analysis = strip_thinking(response.message.content)

        for col in [1, 2]:
            cell           = ws.cell(row=current_row, column=col)
            cell.font      = Font(bold=True, color=HEADER_FONT_COLOR, size=12)
            cell.fill      = PatternFill("solid", fgColor="2E75B6")
            cell.alignment = Alignment(vertical="center")
        ws.cell(row=current_row, column=1, value="Table")
        ws.cell(row=current_row, column=2, value=f"{label}  [{source}]")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        ws.cell(row=current_row, column=1, value="Dimensions").font = Font(bold=True)
        ws.cell(row=current_row, column=2,
                value=f"{tbl['rows']} rows × {tbl['cols']} cols | {source}")
        current_row += 1

        ws.cell(row=current_row, column=1, value="Analysis").font = Font(bold=True)
        cell           = ws.cell(row=current_row, column=2, value=analysis)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill      = PatternFill("solid", fgColor=SUMMARY_BG_COLOR)
        ws.row_dimensions[current_row].height = max(
            10, analysis.count("\n") + len(analysis) // 100
        ) * 15
        current_row += 2

    wb.save(out_path)
    print(f"  ✅ Analysis sheet written")


# ── MAIN ──────────────────────────────────────────────────────────────────

def main(pdf_path: str):
    print("=" * 60)
    print("  PDF Table Extractor → Excel")
    print(f"  OCR Model:  {OCR_MODEL}")
    print(f"  LLM Model:  {TEXT_MODEL}")
    print(f"  Source:     {pdf_path}")
    print("=" * 60)

    tables = extract_tables_from_pdf(pdf_path)

    if not tables:
        print("\n❌ No tables found in this PDF.")
        return

    out_path = export_to_excel(tables, pdf_path)

    native_n = sum(1 for t in tables if t.get("source") == "native")
    image_n  = sum(1 for t in tables if t.get("source") == "image")

    print(f"\n✅ Export complete!")
    print(f"   File:    {out_path}")
    print(f"   Sheets:  Contents | {' | '.join(t['label'] for t in tables)}")
    print(f"   Native tables:      {native_n}")
    print(f"   Image-based tables: {image_n}")

    print(f"\n{'─'*60}")
    print(f"  AI Analysis ({TEXT_MODEL}) adds per table:")
    print(f"    • What the table represents")
    print(f"    • Key patterns and outliers")
    print(f"    • Data quality observations")
    print(f"    • Business insights")
    print(f"{'─'*60}")

    while True:
        choice = input(
            "\nWould you like AI analysis added to the Excel file? [y/n]: "
        ).strip().lower()
        if choice in ("y", "yes"):
            analyse_tables(tables, out_path)
            print(f"\n🎉 Done!")
            print(f"   File:   {out_path}")
            print(f"   Sheets: Analysis | Contents | "
                  f"{' | '.join(t['label'] for t in tables)}")
            break
        elif choice in ("n", "no"):
            print(f"\n🎉 Done!")
            print(f"   File:   {out_path}")
            break
        else:
            print("  Please enter y or n")


if __name__ == "__main__":
    import sys
    pdf = sys.argv[1] if len(sys.argv) > 1 else "./pdfs/test.pdf"
    main(pdf)